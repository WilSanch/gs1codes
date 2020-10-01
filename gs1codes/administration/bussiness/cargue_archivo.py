import openpyxl
from administration.models import Enterprise, Prefix, Code, Range
from administration.bussiness.prefix import prefix_inactivation
from administration.bussiness.models import ErrorList
from administration.common.constants import StCodes
import os.path
from django.db import transaction
import datetime

def valida_excel(excel_file_upload):
    filename = excel_file_upload.name
    nombre, extension = os.path.splitext(filename)

    if  not (extension == ".xls" or extension == ".xlsx"):
        msj= 'Extension no valida ' + str(extension) +  ' las extensiones permitidas son .xls o .xlsx' 
        return msj
    else:
        return ""

def get_container_name(opc):
    if opc == 1:
        container_name = 'devolucion-codigos-test'
    elif opc == 2:
        container_name = 'inactivacion-masiva-prefijos'
    else:
        container_name = 'Null'

    return container_name

def inactivacion_masiva(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb["Datos"]
    
    error_list = []

    i = 0
    for row in worksheet.iter_rows():
        error_list_obj = ErrorList()
        if i == 0:
            i += 1
            continue
        i += 1

        id_prefix = str(row[0].value)
        id_enterprise = str(row[1].value)
        validity_date = row[2].value

        if (id_prefix == None or id_prefix == 'None'):
            error_list_obj["IdRegistro"] = i
            error_list_obj["MsgError"] = "NULL. No se encontraron datos en el registro."
            error_list.append(error_list_obj)
            continue

        enterprise_obj = Enterprise.objects.filter(identification=id_enterprise).first()
        if (not enterprise_obj):
            error_list_obj["IdRegistro"] = i
            error_list_obj["MsgError"] = "ERROR. No se encuentra la empresa'" + id_enterprise + "'."
            error_list.append(error_list_obj)
            continue

        prefix_obj = Prefix.objects.filter(id_prefix=id_prefix,enterprise_id=enterprise_obj.id).first()
        if (not prefix_obj):
            error_list_obj["IdRegistro"] = i
            error_list_obj["MsgError"] = "ERROR. No se encuentra el prefijo '" + id_prefix + "' asignado a la empresa'" + id_enterprise + "'."
            error_list.append(error_list_obj)
            continue

        if (prefix_obj.state_id == StCodes.Suspendido.value):
            error_list_obj["IdRegistro"] = i
            error_list_obj["MsgError"] = "MSG. El prefijo '" + id_prefix + "' se encuentra inactivo.'" + id_enterprise + "'."
            error_list.append(error_list_obj)
            continue

        fecha = datetime.datetime.today()
        with transaction.atomic():
            resp = prefix_inactivation(prefix_obj,fecha,"INACTIVACIÓN MASIVA","")
        if (resp == ""):
            resp = "CORRECTO. Se inactivó el prefijo: '" + id_prefix + "'."

        error_list_obj["IdRegistro"] = i
        error_list_obj["MsgError"] = resp
        error_list.append(error_list_obj)

    return error_list

def devolucion_masiva(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb["Datos"]
    
    error_list = []

    i = 0
    for row in worksheet.iter_rows():
        error_list_obj = ErrorList()
        if i == 0:
            i += 1
            continue
        i += 1

        id_prefix = str(row[0].value)
        id_enterprise = str(row[1].value)

        if (id_prefix == None or id_prefix == 'None'):
            error_list_obj["IdRegistro"] = i
            error_list_obj["MsgError"] = "NULL. No se encontraron datos en el registro."
            error_list.append(error_list_obj)
            continue

        enterprise_obj = Enterprise.objects.filter(identification=id_enterprise).first()
        if (not enterprise_obj):
            error_list_obj["IdRegistro"] = i
            error_list_obj["MsgError"] = "ERROR. No se encuentra la empresa '" + id_enterprise + "'."
            error_list.append(error_list_obj)
            continue

        prefix_obj = Prefix.objects.filter(id_prefix=id_prefix,enterprise_id=enterprise_obj.id).first()
        if (not prefix_obj):
            error_list_obj["IdRegistro"] = i
            error_list_obj["MsgError"] = "ERROR. No se encuentra el prefijo '" + id_prefix + "' asignado a la empresa'" + id_enterprise + "'."
            error_list.append(error_list_obj)
            continue

        with transaction.atomic():
            codes_count = Code.objects.filter(prefix_id= prefix_obj.id).count()
            codes = Code.objects.filter(prefix_id= prefix_obj.id).filter(state_id=1).delete()

            range_obj = Range.objects.filter(id=prefix_obj.range_id).first()

            enterprise_obj.code_quantity_reserved += range_obj.quantity_code - codes_count
            enterprise_obj.code_residue = enterprise_obj.code_quantity_purchased - enterprise_obj.code_quantity_reserved
            prefix_obj.enterprise_id = None
            prefix_obj.save()
            enterprise_obj.save()

        resp = "CORRECTO. Devolución del prefijo: '" + id_prefix + "' y los códigos asignados a este."

        error_list_obj["IdRegistro"] = i
        error_list_obj["MsgError"] = resp
        error_list.append(error_list_obj)

    return error_list